import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.worksheet.datavalidation import DataValidation

def set_column_width_and_wrap(file_path, encoding='utf-8-sig'):
    """
    设置Excel文件中所有sheet的列宽、自动换行、添加新列并设置数据验证
    
    Args:
        file_path (str): Excel文件路径
        encoding (str): 文件编码，默认为utf-8-sig
    """
    try:
        # 打开Excel文件
        workbook = openpyxl.load_workbook(file_path)
        
        # 定义列宽设置（现在有6列）
        column_widths = {
            'A': 12,  # 第一列
            'B': 30,  # 第二列
            'C': 65,  # 第三列
            'D': 30,  # 第四列
            'E': 15,  # 第五列（新增）- 稍微加宽以适应新表头
            'F': 15   # 第六列（新增）- 稍微加宽以适应新表头
        }
        
        # 遍历所有工作表
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            print(f"正在处理工作表: {sheet_name}")
            
            # 设置表头样式（加粗、居中）
            header_font = Font(bold=True)
            header_alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            
            # 获取第一行的列数，确保处理所有现有表头
            max_col_with_data = worksheet.max_column if worksheet.max_column else 4
            
            # 为现有表头设置加粗居中样式
            for col in range(1, max(max_col_with_data + 1, 7)):  # 确保至少处理到F列
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.alignment = header_alignment
            
            # 在第五列和第六列的第一行添加新表头
            worksheet['E1'] = '操作描述准确'
            worksheet['F1'] = '操作复现结果'
            
            # 设置新表头的格式（加粗居中）
            worksheet['E1'].font = header_font
            worksheet['E1'].alignment = header_alignment
            worksheet['F1'].font = header_font
            worksheet['F1'].alignment = header_alignment
            
            # 设置列宽
            for column, width in column_widths.items():
                worksheet.column_dimensions[column].width = width
            
            # 设置单元格格式
            max_row = max(worksheet.max_row, 100)  # 至少处理100行
            
            for row in range(1, max_row + 1):
                for col in range(1, 7):  # 现在有6列
                    cell = worksheet.cell(row=row, column=col)
                    
                    if row == 1:  # 第一行（表头）
                        # 表头已经在上面设置过了，这里跳过
                        continue
                    else:  # 数据行
                        if col <= 4:  # 前4列左对齐
                            cell.alignment = Alignment(wrap_text=True, 
                                                     vertical='top', 
                                                     horizontal='left')
                        else:  # 第5、6列居中对齐
                            cell.alignment = Alignment(wrap_text=True, 
                                                     vertical='center', 
                                                     horizontal='center')
            
            # 创建第五列的数据验证规则（操作描述准确：准确或模糊）
            data_validation_e = DataValidation(
                type="list",
                formula1='"准确,模糊"',
                allow_blank=True,
                showInputMessage=True,
                promptTitle="输入提示",
                prompt="请选择：准确 或 模糊",
                showErrorMessage=True,
                errorTitle="输入错误",
                error="只能选择'准确'或'模糊'！"
            )
            
            # 创建第六列的数据验证规则（操作复现一致：成功或失败）
            data_validation_f = DataValidation(
                type="list",
                formula1='"成功,失败"',
                allow_blank=True,
                showInputMessage=True,
                promptTitle="输入提示",
                prompt="请选择：成功 或 失败",
                showErrorMessage=True,
                errorTitle="输入错误",
                error="只能选择'成功'或'失败'！"
            )
            
            # 将数据验证应用到E列和F列（从第2行开始，第1行是表头）
            data_validation_e.add(f"E2:E{max_row}")
            data_validation_f.add(f"F2:F{max_row}")
            
            # 添加数据验证到工作表
            worksheet.add_data_validation(data_validation_e)
            worksheet.add_data_validation(data_validation_f)
            
            print(f"  - 已设置所有表头为加粗居中")
            print(f"  - 已添加表头: E1='操作描述准确', F1='操作复现结果'")
            print(f"  - 已设置E2:E{max_row}的数据验证（只能选择：准确/模糊）")
            print(f"  - 已设置F2:F{max_row}的数据验证（只能选择：成功/失败）")
        
        # 保存文件
        workbook.save(file_path)
        print(f"\n文件 {file_path} 处理完成！")
        print("列宽设置:")
        for col, width in column_widths.items():
            print(f"  列{col}: {width}")
        
        print("\n格式设置:")
        print("- 所有表头都已设置为加粗居中")
        print("- 所有列都已设置自动换行")
        print("- E列数据验证：只能选择'准确'或'模糊'")
        print("- F列数据验证：只能选择'成功'或'失败'")
        
    except FileNotFoundError:
        print(f"错误: 找不到文件 {file_path}")
    except Exception as e:
        print(f"处理文件时出现错误: {str(e)}")

def main():
    """
    主函数
    """
    # 使用原始字符串避免转义字符问题
    file_path = r"cherry-docs-modules\cherry-docs-modules.xlsx"
    
    # 检查文件是否存在
    import os
    if not os.path.exists(file_path):
        print(f"错误: 文件 {file_path} 不存在，请检查文件路径")
        return
    
    print(f"开始处理文件: {file_path}")
    set_column_width_and_wrap(file_path, encoding='utf-8-sig')

if __name__ == "__main__":
    main()
